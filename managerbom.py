import re
import time
import pprint
import operator
import openpyxl

from tqdm import tqdm
from openpyxl.styles import Font

import inventorapi as api


class ManagerBOM:
    _ROW_PROPERTIES = ("Quantity", "Item Quantity")
    _PURCHASED_ITEM_PROPERTIES = (
        "Part Number",
        "Description",
        "Project",
        "Vendor",
        "Stock Number",
    )
    _MATERIAL_ITEM_PROPERTIES = ("Part Number", "Description", "Material", "Mass")
    _FLAT_MATERIAL_PROPERTIES = (
        "Flat Pattern Width",
        "Flat Pattern Length",
        "Flat Pattern Area",
    )

    def __init__(self, template: openpyxl.Workbook, iam_inventor_api_object):
        self._template = template
        self._iam = api.Assembly(iam_inventor_api_object)
        self._bom_view = self._iam.raw_bom
        self._data = []
        self._exceptions = []

    @property
    def data(self):
        return self._data

    @property
    def exceptions(self):
        return self._exceptions

    @property
    def bom_biew(self):
        return self._bom_view

    def _collect_data(self, rows: list[api.BOMRow], multiplier: int = 1):
        # Base case for the recursion
        if not rows:
            return

        for row in tqdm(rows):
            item_quantity = row.item_quantity * multiplier
            # Common data for all types of items
            row_data = {
                "Unit Quantity": row.item.unit_quantity,
                "Item Quantity": item_quantity,
            }
            initial_row_data_size = len(row_data)

            # Data for purchased items
            if row.is_purchased():
                row_data.update(
                    row.item.get_properties(ManagerBOM._PURCHASED_ITEM_PROPERTIES)
                )
                row_data["Type"] = "purchased"

            else:
                # Start of a recursion, look for child rows of the row
                self._collect_data(row.get_child_rows(), item_quantity)

                # Common data for material items
                if row.is_normal() and row.item.is_part():
                    row_data.update(
                        row.item.get_properties(ManagerBOM._MATERIAL_ITEM_PROPERTIES)
                    )

                    # Data for profile material items
                    if row.item.is_modeling():
                        row_data.update(row.item.get_size())
                        row_data["Type"] = "profile material"

                    # Data for sheet material items
                    elif row.item.is_sheet_metal():
                        # Consider strips as profile materials
                        if row_data["Material"].startswith("Полоса"):
                            size = row.item.get_properties(
                                ("Flat Pattern Length", "Flat Pattern Width")
                            )
                            # Convert cm to mm
                            row_data["Size Z"] = max(size.values()) * 10
                            row_data["Size X"] = min(size.values()) * 10
                            # Thickness of the strips, relatively small
                            row_data["Size Y"] = 0
                            row_data["Type"] = "profile material"

                        else:
                            row_data.update(
                                row.item.get_properties(
                                    ManagerBOM._FLAT_MATERIAL_PROPERTIES
                                )
                            )
                            row_data["Type"] = "sheet material"

            if len(row_data) > initial_row_data_size:
                self._data.append(row_data)

    def _summarize_purchased(self):
        """
        Summarize data in a list of purchased items.
        Each item as a list.
        item[0]: Description, if Vendor or Project, else Part Number
        item[1]: Project
        item[2]: Vendor
        item[3]: Quantity(int or str for measured items)
        item[4]: Stock Number
        """
        print("Summarizing purchased items")
        purchased = []

        for item in filter(lambda i: i["Type"] == "purchased", self._data):
            project = item["Project"]
            vendor = item["Vendor"]
            stock = item["Stock Number"]

            # Calculate quantity
            try:
                unit_quantity = float(item["Unit Quantity"].rstrip(" м"))
            except ValueError:
                unit_quantity = 1
            finally:
                quantity = unit_quantity * item["Item Quantity"]

            # Choose correct value for description
            if project or vendor:
                description = item["Description"]
            else:
                description = item["Part Number"]

            new_item = [description, project, vendor, quantity, stock]

            to_append = True
            for item in purchased:
                # Compare discription of new item with over items
                if new_item[0] == item[0]:
                    # Update quantity of already existing item
                    item[3] += new_item[3]
                    to_append = False

            if to_append:
                purchased.append(new_item)

        # Add units for quantity for measured items
        for item in purchased:
            if isinstance(item[3], float):
                item[3] = f"{item[3]:.2f} м"

        # Sort by vendor and description
        purchased.sort(key=operator.itemgetter(2, 0))
        print("Purchased items: ", len(purchased))
        return purchased

    def _summarize_unified(self):
        """
        Summarize data in a list of unified items. Each item as a list.
        item[0]: Part Number
        item[1]: Description
        item[2]: Quantity
        """
        print("Summarizing unified items")
        unified = []

        for item in filter(lambda i: i["Part Number"].startswith("МД1000"), self._data):
            part_number = item["Part Number"]
            description = item["Description"]
            quantity = item["Item Quantity"]

            new_item = [part_number, description, quantity]

            to_append = True
            for item in unified:
                # Compare part number of new item with over items
                if new_item[0] == item[0]:
                    # Update quantity of already existing item
                    item[2] += new_item[2]
                    to_append = False

            if to_append:
                unified.append(new_item)

        # Sort by vendor and description
        unified.sort(key=lambda i: i[0])
        print("Unified items: ", len(unified))
        return unified

    def _summarize_sheet_material(self):
        """
        Summarize data in a list of sheet materials.
        Each material as a list.
        item[0]: Material
        item[1]: Mass
        item[2]: Area
        """
        print("Summarizing sheet material items")
        sheet_material = []

        for item in filter(lambda i: i["Type"] == "sheet material", self._data):
            quantity = item["Item Quantity"]
            material = item["Material"]
            # Convert mass from grams to kilograms
            mass = item["Mass"] / 1000 * quantity
            # Convert area from cm2 to m2
            area = item["Flat Pattern Area"] / 10000 * quantity

            new_item = [material, mass, area]

            to_append = True
            for item in sheet_material:
                if new_item[0] == item[0]:
                    item[1] += new_item[1]
                    item[2] += new_item[2]
                    to_append = False

            if to_append:
                sheet_material.append(new_item)

        for material in sheet_material:
            item[1] = round(item[1], 2)
            item[2] = round(item[2], 2)

        print("Sheet material items: ", len(sheet_material))
        return sheet_material

    def _summarize_profile_material(self):
        """
        Summarize data in a list of profile materials.
        Each material as a list.
        item[0]: Material
        item[1]: Mass
        item[2]: Length
        """
        print("Summarizing profile material items")
        profile_material = []

        for item in filter(lambda i: i["Type"] == "profile material", self._data):
            quantity = item["Item Quantity"]
            material = item["Material"]
            # Convert mass from grams to kilograms
            mass = item["Mass"] / 1000 * quantity

            # Deside which size of the item use as length
            # based on profile size from material name
            matches = re.match(r"(\d+).(\d+)?", material)
            try:
                profile_size = matches.groups()
            except AttributeError:
                profile_size = ""
            finally:
                # Primarily consider size along Z axis
                axis = "Z"
                if str(item[f"Size {axis}"]) in profile_size:
                    while True:
                        axis = input(
                            f"Axis of length for {item['Part Number']} (X, Y or Z): "
                        ).upper()
                        if axis in ("X", "Y", "Z"):
                            break

                # Convert length in millimetres to metres
                length = item[f"Size {axis}"] / 1000 * quantity

            new_item = [material, mass, length]

            to_append = True
            for item in profile_material:
                if new_item[0] == item[0]:
                    item[1] += new_item[1]
                    item[2] += new_item[2]
                    to_append = False

            if to_append:
                profile_material.append(new_item)

        print("Profile material items: ", len(profile_material))
        return profile_material

    def _transfer_purchased(self):
        sheet = self._template[self._template.sheetnames[2]]
        self._transfer(sheet, self._summarize_purchased())

    def _transfer_unified(self):
        sheet = self._template[self._template.sheetnames[3]]
        self._transfer(sheet, self._summarize_unified())

    def _transfer(self, sheet, data):
        print(f"Transferring data to {sheet.title}")
        start_row = 3
        for i, item in enumerate(data):
            row = start_row + i
            sheet[f"A{row}"] = i + 1
            for j, prop in enumerate(item):
                column = j + 2
                sheet.cell(column=column, row=row).value = prop

    def _transfer_profile_material(self):
        sheet = self._template[self._template.sheetnames[0]]
        self._transfer_material(sheet, self._summarize_profile_material())

    def _transfer_sheet_material(self):
        sheet = self._template[self._template.sheetnames[1]]
        self._transfer_material(sheet, self._summarize_sheet_material())

    def _transfer_material(self, sheet, data):
        print(f"Transferring data to {sheet.title}")
        # Font to highlight cells in changed rows
        highlight = Font(color="FF0000")

        for item in data:
            excepted_item = True
            for cell in sheet["J"]:
                # Find the target row of material
                if cell.value and item[0].startswith(cell.value):
                    excepted_item = False
                    row = cell.row
                    # Higlight the row
                    for c in sheet[row]:
                        c.font = highlight
                    # Trasfer data
                    sheet[f"D{row}"] = item[1]
                    sheet[f"G{row}"] = item[2]

            if excepted_item:
                self._exceptions.append(item)

    def issue_bom(
        self, profile_material=True, sheet_material=True, purchased=True, unified=True
    ) -> openpyxl.Workbook:
        print("Collecting data")
        self._collect_data(self._bom_view.get_rows())
        print(f"Data size is {len(self._data)}")
        print()

        if profile_material:
            self._transfer_profile_material()
            print()
            time.sleep(1)

        if sheet_material:
            self._transfer_sheet_material()
            print()
            time.sleep(1)

        if purchased:
            self._transfer_purchased()
            print()
            time.sleep(1)

        if unified:
            self._transfer_unified()
            print()
            time.sleep(1)

        # Print not transferred data
        if self._exceptions:
            print("Could not transfer:")
            pprint.pprint(self._exceptions)

        return self._template
